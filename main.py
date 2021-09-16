import builtins
import os
import shutil
import time
import openpyxl
from bs4 import BeautifulSoup

nexus_group_code = {}
nexus_alias = {}


def read_reporting(root, file, workbook, total_workbook):
    date_str = time.strftime("%Y年%m月%d日", time.localtime())
    xml_path = root + '/' + file
    sheet_names = workbook.sheetnames
    worksheet = workbook.active
    total_worksheet = total_workbook.active
    soup = BeautifulSoup(builtins.open(xml_path, encoding='utf8'), 'xml')
    problems = soup.find_all(name='problem')
    # 追加缺陷记录
    for problem in problems:
        # 模块
        v_module = problem.module.text.upper()
        # 缺陷位置
        v_location = problem.entry_point['FQNAME'].split()
        # 行号
        v_line = problem.line.text
        # 问题分类
        v_problem_class = problem.problem_class.text
        # 严重程度
        v_severity = problem.problem_class['severity']
        # 判断默认优先级
        v_priority = 'Immediate'
        # 默认状态为Open
        v_status = 'Open'
        if v_severity == 'BLOCKER':
            v_severity = 'Blocker'
            v_priority = 'Urgent'
        elif v_severity == 'CRITICAL':
            v_severity = 'Critical'
            v_priority = 'High'
        elif v_severity == 'MAJOR':
            v_severity = 'Major'
            v_priority = 'Normal'
        elif v_severity == 'MINOR':
            v_severity = 'Minor'
            v_priority = 'Low'
        # 缺陷描述
        v_description = problem.description.text
        # 获取真实模块名
        if v_module in nexus_alias:
            v_module = nexus_alias[v_module]
        # 查询开发团队
        develop_group_name = None
        if v_module in nexus_group_code:
            develop_group_name = nexus_group_code[v_module]
        # 追加新列
        worksheet.append(
            [None, date_str, develop_group_name, v_module, v_problem_class, v_location[0] + ':' + v_line, v_description,
             v_severity, v_priority, v_status])
        total_worksheet.append(
            [None, date_str, develop_group_name, v_module, v_problem_class, v_location[0] + ':' + v_line, v_description,
             v_severity, v_priority, v_status])


def read_nexus():
    workbook = openpyxl.load_workbook("./nexus.xlsx")
    worksheet = workbook['code']
    for row in worksheet.rows:
        nexus_group_code[row[2].value] = row[1].value
    worksheet = workbook['alias']
    for row in worksheet.rows:
        nexus_alias[row[0].value] = row[1].value
    workbook.close()


def aim(projects_dir):
    # 读取团队与模块的对应关系
    read_nexus()

    # 删除已存在xlsx，复制模版到其位置
    for root, dirs, files in os.walk(projects_dir):
        if len(dirs) != 0:
            continue
        # 删除原有xlsx
        for file in files:
            if file.endswith('.xlsx'):
                os.remove(root + '/' + file)
        # 复制模版
        project_name = root.split('/')[-1]
        shutil.copyfile('./template.xlsx', './xlsx_reporting/' + project_name + '.xlsx')

    # 复制汇总表
    total_xlsx_path = './xlsx_reporting/total.xlsx'
    shutil.copyfile('./template.xlsx', total_xlsx_path)
    total_workbook = openpyxl.load_workbook(total_xlsx_path)

    # 开始写入xlsx
    for root, dirs, files in os.walk(projects_dir):
        if len(dirs) != 0:
            continue
        for file in files:
            if file.startswith('Ali') & file.endswith('.xml'):
                xml_path = root + '/' + file
                # 打开对应的excel
                project_name = root.split('/')[-1]
                xlsx_path = './xlsx_reporting/' + project_name + '.xlsx'
                workbook = openpyxl.load_workbook(xlsx_path)
                read_reporting(root, file, workbook, total_workbook)
                print('[Conversion done] {0} --> {1}'.format(xml_path, xlsx_path))
                workbook.save(xlsx_path)
                workbook.close()
    total_workbook.save(total_xlsx_path)
    total_workbook.close()


if __name__ == '__main__':
    aim('xml_reporting')
